import os
import shutil
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  
SRC_DIR = os.path.join(BASE_DIR, "src")
RESULTS_DIR = os.path.join(BASE_DIR, "results")
TEMPLATE_PATH = os.path.join(SRC_DIR, "Template.xlsx")


os.makedirs(RESULTS_DIR, exist_ok=True)
txt_files = [f for f in os.listdir(SRC_DIR) if f.lower().endswith(".txt")]

SHEETS_TO_CLEAN = {
    "CONCENTRATED": 3,
    "L3OUTS": 3,
    "PG": 3,
    "LEAF_INFO": 3,
    "VPCs": 4,      
    "VPC": 3,
    "PORTs": 3,
    "Statics": 3,
    "PORTS_": 3
}

def clean_excel(file_path):
    try:
        wb = load_workbook(file_path)

        for sheet_name, start_row in SHEETS_TO_CLEAN.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_row = ws.max_row

                if max_row >= start_row:
                    ws.delete_rows(start_row, max_row - start_row + 1)
                    #print(f"🧹 {sheet_name}: filas {start_row}–{max_row} eliminadas")

        wb.save(file_path)
        wb.close()
        print(f"✅ Limpieza completada para {os.path.basename(file_path)}")

    except Exception as e:
        print(f"❌ Error limpiando {file_path}: {e}")

def dupplicate():
    if not txt_files:
        print("⚠️ No se encontraron archivos .txt en:", SRC_DIR)
    else:
        for txt_file in txt_files:
            name_part = os.path.splitext(txt_file)[0][:17]
            dest_file = os.path.join(RESULTS_DIR, f"{name_part}.xlsx")
            shutil.copy2(TEMPLATE_PATH, dest_file)
            #print(f"📄 Copiado: {txt_file} → {dest_file}")
            clean_excel(dest_file)

        print(f"\n🎉 {len(txt_files)} archivos .xlsx creados y limpiados en {RESULTS_DIR}")


# --- Ejecutar ---
if __name__ == "__main__":
    dupplicate()
