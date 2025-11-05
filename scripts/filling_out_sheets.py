import os
import re
from openpyxl import load_workbook, Workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  
SRC_DIR = os.path.join(BASE_DIR, "src")
RESULTS_DIR = os.path.join(BASE_DIR, "results")
TEMPLATE_PATH = os.path.join(SRC_DIR, "Template.xlsx")
RESUMEN_PATH = os.path.join(RESULTS_DIR, "resumen.xlsx")

os.makedirs(RESULTS_DIR, exist_ok=True)
txt_files = [f for f in os.listdir(SRC_DIR) if f.lower().endswith(".txt")]


def update_resumen_excel(combined, txt_name):
    """Crea o actualiza el Excel 'resumen.xlsx' con la información procesada."""
    resumen_headers = [ "REGION", "FABRIC_NAME", "POD", "NODE", "ADDRESS", "VERSION", "SERIAL", "NAME", "MODEL"]
    
    # Obtener el primer dígito del nombre del archivo
    import re
    match = re.search(r"\d", txt_name)
    first_digit = match.group(0) if match else ""

    # Abrir o crear el Excel resumen
    if os.path.exists(RESUMEN_PATH):
        wb_resumen = load_workbook(RESUMEN_PATH)
        ws = wb_resumen.active
    else:
        wb_resumen = Workbook()
        ws = wb_resumen.active
        ws.title = "Resumen"
        ws.append(resumen_headers)

    # Crear índice de nombres existentes
    name_to_row = {str(ws[f"G{row}"].value): row for row in range(2, ws.max_row + 1) if ws[f"G{row}"].value}

    for item in combined:
        name = item.get("name", "")
        pod = item.get("pod", "")
        node = item.get("node", "")
        address = item.get("address", "")
        version = item.get("version", "")
        serial = item.get("serial", "")
        model = item.get("model", "")

        if not name:
            continue

        # Si ya existe el nombre → sobreescribir
        if name in name_to_row:
            row = name_to_row[name]
        else:
            row = ws.max_row + 1

        ws[f"A{row}"] = first_digit
        ws[f"B{row}"] = txt_name
        ws[f"C{row}"] = pod
        ws[f"D{row}"] = node
        ws[f"E{row}"] = address
        ws[f"F{row}"] = version
        ws[f"G{row}"] = serial
        ws[f"H{row}"] = name
        ws[f"I{row}"] = model

        

    wb_resumen.save(RESUMEN_PATH)

def fill_ports_sheet(all_data):
    for site, data in all_data.items():
        combined = data.get("l1PhysIf_ethpmPhysIf", [])
        if not combined:
            print(f"⚠️ {site}: No hay datos combinados, se omite Excel")
            continue

        excel_name = f"{site[:17]}.xlsx"
        excel_path = os.path.join(RESULTS_DIR, excel_name)

        if not os.path.exists(excel_path):
            print(f"⚠️ No se encontró {excel_name} en {RESULTS_DIR}")
            continue

        wb = load_workbook(excel_path)
        if "PORTs" not in wb.sheetnames or "PORTS_" not in wb.sheetnames:
            print(f"⚠️ {excel_name} no tiene hojas requeridas (PORTs / PORTS_)")
            wb.close()
            continue

        # --- PORTs ---
        ws_ports = wb["PORTs"]

        # Limpiar desde la fila 2
        for row in ws_ports.iter_rows(min_row=2, max_col=6):
            for cell in row:
                cell.value = None

        # Escribir los nuevos valores
        for i, item in enumerate(combined, start=2):
            ws_ports[f"A{i}"] = item.get("adminSt", "")
            ws_ports[f"B{i}"] = item.get("operSt", "")
            ws_ports[f"C{i}"] = item.get("descr", "")
            ws_ports[f"D{i}"] = item.get("pod", "")
            ws_ports[f"E{i}"] = item.get("node", "")
            ws_ports[f"F{i}"] = item.get("phys", "")

        # --- PORTS_ ---
        ws_ports_ = wb["PORTS_"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"F{i}"] = item.get("operSt", "")
            ws_ports_[f"G{i}"] = item.get("operStQual", "")
            ws_ports_[f"I{i}"] = item.get("dn_2", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["A", "B", "C", "H", "J"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\D)2\b', lambda m: f"{m.group(1)}{i}", formula)
                    ws_ports_[f"{col}{i}"] = new_formula

        ws_ports_ = wb["CONCENTRATED"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"A{i}"] = item.get("node", "")
            ws_ports_[f"B{i}"] = item.get("pod", "")
            ws_ports_[f"D{i}"] = item.get("phys", "")
            ws_ports_[f"E{i}"] = item.get("adminSt", "")
            ws_ports_[f"F{i}"] = item.get("operSt", "")
            ws_ports_[f"H{i}"] = item.get("descr", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["C", "G", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\$?[A-Z]{1,3}\$?)2\b', lambda m: f"{m.group(1)}{i}", formula)

                    ws_ports_[f"{col}{i}"] = new_formula

        combined = data.get("eqptCh_topSystem", [])
        if not combined:
            print(f"⚠️ {site}: No hay datos combinados, se omite Excel")
            continue
        # --- PORTS_ ---
        ws_ports_ = wb["LEAF_INFO"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"A{i}"] = int(item.get("pod", ""))
            ws_ports_[f"B{i}"] = int(item.get("node", ""))
            ws_ports_[f"D{i}"] = item.get("address", "")
            ws_ports_[f"E{i}"] = item.get("version", "")
            ws_ports_[f"F{i}"] = item.get("serial", "")
            ws_ports_[f"G{i}"] = item.get("name", "")
            ws_ports_[f"H{i}"] = item.get("model", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["C"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\D)2\b', lambda m: f"{m.group(1)}{i}", formula)
                    ws_ports_[f"{col}{i}"] = new_formula

        update_resumen_excel(combined, site)




        combined = data.get("l3extRsPathL3OutAtt", [])
        # --- PORTS_ ---
        ws_ports_ = wb["L3OUTS"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"A{i}"] = item.get("dn", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["B", "C", "D", "E", "F", "G", "H"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\D)2\b', lambda m: f"{m.group(1)}{i}", formula)
                    ws_ports_[f"{col}{i}"] = new_formula




        combined = data.get("infraRsAccBaseGrp", [])
        # --- PORTS_ ---
        ws_ports_ = wb["PG"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"D{i}"] = item.get("dn", "")
            ws_ports_[f"G{i}"] = item.get("tDn", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["A", "B", "C", "E", "F", "H"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\D)2\b', lambda m: f"{m.group(1)}{i}", formula)
                    ws_ports_[f"{col}{i}"] = new_formula



        combined = data.get("fvRsPathAtt", [])
        # --- PORTS_ ---
        ws_ports_ = wb["Statics"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"A{i}"] = item.get("dn", "")
            ws_ports_[f"I{i}"] = item.get("encap", "")
            ws_ports_[f"J{i}"] = item.get("mode", "")

        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["B", "C", "D", "E", "F", "G", "H"]
        formula_templates = {}

        # Guardar las fórmulas originales
        for col in formula_cols:
            cell = ws_ports_[f"{col}2"]
            if cell.data_type == "f" or isinstance(cell.value, str) and cell.value.startswith("="):
                formula_templates[col] = cell.value
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente
        for i in range(3, len(combined) + 2):  # Desde la fila 3 hasta última
            for col, formula in formula_templates.items():
                if formula:
                    # Reemplazar todas las referencias numéricas de fila (e.g. I2 → I3)
                    new_formula = re.sub(r'(\D)2\b', lambda m: f"{m.group(1)}{i}", formula)
                    ws_ports_[f"{col}{i}"] = new_formula


        combined = data.get("fabricExplicitGEp", [])
        # --- PORTS_ ---
        ws_ports_ = wb["VPC"]

        # Escribir datos base
        for i, item in enumerate(combined, start=2):
            ws_ports_[f"A{i}"] = item.get("name", "")
            ws_ports_[f"B{i}"] = item.get("id", "")

        
        ws_ports_ = wb["VPCs"]
        unique_names = list({item.get("name", "") for item in combined})
        for i, name in enumerate(unique_names, start=3):
            ws_ports_[f"A{i}"] = name
        
        # Copiar fórmulas de la fila 2 (A, B, C, H, J)
        formula_cols = ["B", "C", "D", "E", "F", "G", "H", "I", "J",
                "K", "L", "M", "N", "O", "P", "Q", "R", "S"]
        formula_templates = {}

        # Guardar fórmulas de la fila 3
        for col in formula_cols:
            cell = ws_ports_[f"{col}3"]
            val = cell.value
            if isinstance(val, str) and val.startswith("="):  # Solo si es fórmula
                formula_templates[col] = val
            else:
                formula_templates[col] = None

        # Replicar fórmulas dinámicamente desde fila 4
        for i in range(4, 3 + len(unique_names)):
            for col, formula in formula_templates.items():
                if not formula:
                    continue  # Saltar si no hay fórmula

                # Eliminar saltos de línea o tabulaciones internas
                cleaned_formula = re.sub(r'\s+', ' ', formula)

                # Reemplazar referencias de fila 3 -> fila actual
                # Captura expresiones como A3, $A3, AA3, pero no números sueltos
                new_formula = re.sub(
                    r'(?<![A-Z0-9_$])(\$?[A-Z]{1,3})3(?![0-9])',
                    lambda m: f"{m.group(1)}{i}",
                    cleaned_formula
                )

                ws_ports_[f"{col}{i}"] = new_formula




        wb.save(excel_path)
        wb.close()


